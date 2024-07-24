import requests
import json
import openpyxl
from openpyxl import Workbook

class StockAnalyzer:
    def __init__(self, config_file= r'old/config.json'):
        self.workbook = Workbook()
        self.results_sheet = self.workbook.active
        self.results_sheet.title = "Results"
        self.yahoo_api_key = ""
        self.fidelity_api_key = ""

        # Load configuration
        config = None
        try: 
            with open(config_file, 'r') as f:
                config = json.load(f)
        except FileNotFoundError:
            print(f"File not found: {config_file}")
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
    
        if config is None:
            print("Config not loaded")
        else:
            print(config)

        if config is not None:
            self.yahoo_api_key = config.get('YahooApiKey', '')
            self.fidelity_api_key = config.get('FidelityApiKey', '')
            self.tickers = config.get('Tickers',[])
        else:
            print("Config not loaded.  Using default values.")
            self.yahoo_api_key = 'e2b2dc5f7emshe1fd4861fb4c1d0p1716fejsnb82cd7419bd1'
            self.fidelity_api_key = 'e2b2dc5f7emshe1fd4861fb4c1d0p1716fejsnb82cd7419bd1'
            self.tickers = ["AAPL", "GOOGL", "MSFT", "AMZN", "CRWD", "NU", "NXT"]


    def analyze_stocks(self):
        try:
            # Clear previous results
            self.results_sheet.delete_rows(1, self.results_sheet.max_row)
            
            #set up headers in Results Sheets 
            self.setup_results_header() 

            #Loop through tickers and fetch data
            for ticker in self.tickers:
                self.fetch_and_process_stock_data(ticker)

            # Calculate buy ratings 
            self.calculate_stock_buy_ratings_for_analysis() 

            print("Stock analysis complete!")            
        except Exception as e:
            print(f"Error in analyze_stocks: {str(e)}")

    def get_config_value(self, key):
        for row in self.config_sheet.iter_rows(values_only=True):
            if row[0] == key:
                return row[1]
        return None

    def get_tickers(self):
        tickers = []
        for cell in self.config_sheet['D']:
            if cell.value:
                tickers.append(cell.value)
        return tickers

    def setup_results_header(self):
        headers = ["Ticker", "Current Price", "PE Ratio", "Dividend Yield", "EPS Recent",
                   "Target Price", "Valuation", "Quality", "Equity Score", "Buy Rating"]
        self.results_sheet.append(headers)

    def fetch_and_process_stock_data(self, ticker):
        yahoo_data = self.fetch_yahoo_data(ticker)
        fidelity_data = self.fetch_fidelity_data(ticker)

        row = [ ticker,
            yahoo_data.get("regularMarketPreviousClose", ""),
            yahoo_data.get("forwardPE", ""),
            yahoo_data.get("dividendRate", ""),
            yahoo_data.get("trailingEPS", ""),
            yahoo_data.get("forwardEPS", ""),
            yahoo_data.get("targetMeanPrice", ""),
            fidelity_data.get("valuation", ""),
            fidelity_data.get("quality", ""),
            fidelity_data.get("equityScore", ""),
        ]
        self.results_sheet.append(row)

    def fetch_yahoo_data(self, ticker):
        url = f"https://apidojo-yahoo-finance-v1.p.rapidapi.com/market/v2/get-quotes?symbols={ticker}"
        headers = {
            "x-rapidapi-key": self.yahoo_api_key,
            "x-rapidapi-host": "apidojo-yahoo-finance-v1.p.rapidapi.com"
        }
        response = requests.get(url, headers=headers)
        data = response.json()

        if "quoteResponse" in data and "result" in data["quoteResponse"] and data["quoteResponse"]["result"]:
            result = data["quoteResponse"]["result"][0]
            return {
                "regularMarketPrice": result.get("regularMarketPrice"),
                "forwardPE": result.get("forwardPE"),
                "dividendYield": result.get("dividendYield"),
                "epsTrailingTwelveMonths": result.get("epsTrailingTwelveMonths"),
                "targetMeanPrice": result.get("targetMeanPrice")
            }
        else:
            print(f"No results found for {ticker}")
            return {}

    def fetch_fidelity_data(self, ticker):
        url = f"https://fidelity-investments.p.rapidapi.com/auto-complete?query={ticker}"
        headers = {
            "x-rapidapi-key": self.fidelity_api_key,
            "x-rapidapi-host": "fidelity-investments.p.rapidapi.com"
        }
        response = requests.get(url, headers=headers)
        data = response.json()

        if data:
            return {
                "valuation": data.get("valuation", 0),
                "quality": data.get("quality", 0),
                "equityScore": data.get("equityScore", 0)
            }
        else:
            print(f"No content returned for {ticker} from Fidelity API")
            return {"valuation": 0, "quality": 0, "equityScore": 0}

    def calculate_stock_buy_ratings_for_analysis(self):
        for row in self.results_sheet.iter_rows(min_row=2, values_only=True):
            eps_score = self.score_eps_change(row[4], 0)  # Assuming 0 for prior EPS
            target_value_score = self.score_target_value(row[1], row[5])
            valuation_score = self.score_valuation(row[6])
            quality_score = self.score_quality(row[7])
            equity_score = self.score_equity_score(row[8])
            pe_score = self.score_pe(row[2])
            dividend_score = self.score_dividend(row[3])

            buy_rating = eps_score + target_value_score + valuation_score + quality_score + equity_score + pe_score + dividend_score
            self.results_sheet.cell(row=self.results_sheet.max_row, column=10, value=buy_rating)

    # Scoring functions
    def score_eps_change(self, eps_recent, eps_prior):
        if eps_recent > eps_prior:
            return 1
        elif eps_recent == eps_prior:
            return 0
        else:
            return -1

    def score_target_value(self, current_price, target_price):
        if current_price > target_price:
            if current_price > target_price * 1.15:
                return -1
            else:
                return 0
        elif current_price >= target_price * 0.85:
            return 1
        elif current_price >= target_price * 0.5:
            return 2
        else:
            return 3

    def score_valuation(self, valuation):
        if 1 <= valuation <= 19:
            return -2
        elif 20 <= valuation <= 39:
            return -1
        elif 40 <= valuation <= 60:
            return 0
        elif 61 <= valuation <= 80:
            return 1
        elif 81 <= valuation <= 99:
            return 2
        else:
            return 0

    def score_quality(self, quality):
        return self.score_valuation(quality)

    def score_equity_score(self, equity_score):
        if equity_score < 3:
            return -1
        elif 3 <= equity_score <= 7:
            return 0
        else:
            return 1

    def score_pe(self, pe):
        if pe < 10:
            return 2
        elif 10 <= pe <= 20:
            return 1
        else:
            return 0

    def score_dividend(self, dividend):
        if dividend < 1:
            return 0
        elif 1 <= dividend <= 10:
            return 2
        else:
            return 3

if __name__ == "__main__":
    analyzer = StockAnalyzer()
    analyzer.analyze_stocks()
    analyzer.workbook.save("stock_analysis_results.xlsx")